#!/usr/bin/env python3
"""
Excel Complexity Analyzer

Analyzes Excel file metadata to determine optimal processing strategy.
Uses lightweight openpyxl scanning without loading full data into memory.

Usage:
    python complexity_analyzer.py <file_path> [sheet_name]

    If sheet_name is not provided, analyzes all sheets.

Output (JSON):
    {
        "file": "<file_path>",
        "sheets": {
            "<sheet_name>": {
                "is_complex": bool,
                "recommended_strategy": "pandas" | "html",
                "reasons": ["reason1", "reason2"],
                "stats": {
                    "total_rows": int,
                    "deep_merges": int,
                    "empty_interruptions": int
                }
            }
        }
    }

Dependencies:
    - openpyxl

Author: Claude Code - excel-parser Skill
Version: 1.0
"""

import openpyxl
from openpyxl.utils import range_boundaries
import json
import sys
from pathlib import Path


class ExcelComplexityAnalyzer:
    """Analyzes Excel file complexity using metadata scanning."""

    def __init__(self, file_path):
        """
        Initialize analyzer with Excel file.

        Args:
            file_path: Path to Excel file (.xlsx, .xlsm)
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load workbook in read-only mode with data_only=True
        # This reads calculated values instead of formulas
        self.wb = openpyxl.load_workbook(
            str(self.file_path),
            read_only=False,  # Need to access merged_cells
            data_only=True
        )

    def analyze_sheet(self, sheet_name):
        """
        Analyze a single sheet's complexity.

        Args:
            sheet_name: Name of the sheet to analyze

        Returns:
            dict: Analysis results with complexity score and recommendation
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = self.wb[sheet_name]

        # Get basic metrics
        max_row = sheet.max_row
        max_col = sheet.max_column
        merged_ranges = sheet.merged_cells.ranges

        # Analyze merged cell distribution
        deep_merges = 0
        shallow_merges = 0

        for merge in merged_ranges:
            # Get bounds of merged range
            min_col, min_row, max_col_merge, max_row_merge = range_boundaries(str(merge))

            # Consider row 5 as boundary between header and data
            if min_row > 5:
                deep_merges += 1
            else:
                shallow_merges += 1

        # Analyze data continuity (check for empty row interruptions)
        empty_interruptions = 0

        # Only check for interruptions in smaller tables to avoid performance issues
        if max_row < 200:
            consecutive_empty = 0
            for row_idx in range(1, max_row + 1):
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

                # Check if entire row is empty
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    consecutive_empty += 1
                else:
                    # Count transition from data to empty to data as interruption
                    if consecutive_empty > 0 and row_idx > 1:
                        empty_interruptions += 1
                    consecutive_empty = 0

        # Apply scoring rules
        is_complex = False
        reasons = []

        # Rule 1: Deep merged cells indicate complex structure
        if deep_merges > 2:
            is_complex = True
            reasons.append(f"Detected {deep_merges} merged cells in data region")

        # Rule 2: Multiple empty row interruptions suggest multi-table layout
        if max_row < 300 and empty_interruptions > 2:
            is_complex = True
            reasons.append(f"Detected {empty_interruptions} empty row interruptions (multi-table layout)")

        # Rule 3: Large tables must use Pandas (HTML would exceed token limits)
        if max_row > 1000:
            is_complex = False
            reasons = [f"Row count ({max_row}) exceeds 1000, forcing Pandas mode"]

        # Rule 4: Default for standard tables
        if not is_complex and not reasons:
            reasons.append("Standard table structure, no complex patterns detected")

        # Determine recommended strategy
        recommended_strategy = "html" if is_complex else "pandas"

        return {
            "is_complex": is_complex,
            "recommended_strategy": recommended_strategy,
            "reasons": reasons,
            "stats": {
                "total_rows": max_row,
                "total_columns": max_col,
                "deep_merges": deep_merges,
                "shallow_merges": shallow_merges,
                "empty_interruptions": empty_interruptions
            }
        }

    def analyze_all(self):
        """
        Analyze all sheets in the workbook.

        Returns:
            dict: Analysis results for all sheets
        """
        results = {
            "file": str(self.file_path),
            "sheets": {}
        }

        for sheet_name in self.wb.sheetnames:
            results["sheets"][sheet_name] = self.analyze_sheet(sheet_name)

        return results

    def close(self):
        """Close the workbook."""
        self.wb.close()


def main():
    """CLI entry point."""
    if len(sys.argv) < 2:
        print("Usage: python complexity_analyzer.py <file_path> [sheet_name]", file=sys.stderr)
        print("\nExample:")
        print("  python complexity_analyzer.py data.xlsx", file=sys.stderr)
        print("  python complexity_analyzer.py data.xlsx Sheet1", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        analyzer = ExcelComplexityAnalyzer(file_path)

        if sheet_name:
            # Analyze single sheet
            result = {
                "file": str(file_path),
                "sheets": {
                    sheet_name: analyzer.analyze_sheet(sheet_name)
                }
            }
        else:
            # Analyze all sheets
            result = analyzer.analyze_all()

        analyzer.close()

        # Output JSON to stdout
        print(json.dumps(result, indent=2, ensure_ascii=False))

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
